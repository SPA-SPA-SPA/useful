#!python3
#coding=utf-8

""" 这个模块分析下载到的网页文件，并写入到一个lxsl文件中 """
import bs4, requests, os
import openpyxl
# openpyxl.utils
from openpyxl.utils import get_column_letter

# 准备一个空列表，存放每个文件中，33个地区的数据
provinces = [
        '全国合计',
        '部本级',
        '北京市',
        '天津市',
        '河北省',
        '山西省',
        '内蒙古自治区',
        '辽宁省',
        '吉林省',
        '黑龙江省',
        '上海市',
        '江苏省',
        '浙江省',
        '安徽省',
        '福建省',
        '江西省',
        '山东省',
        '河南省',
        '湖北省',
        '湖南省',
        '广东省',
        '广西壮族自治区',
        '海南省',
        '重庆市',
        '四川省',
        '贵州省',
        '云南省',
        '西藏自治区',
        '陕西省',
        '甘肃省',
        '青海省',
        '宁夏回族自治区',
        '新疆维吾尔自治区'
    ]

""" 写表行的函数 """
def write_a_row(tds, row_num, path):
    wb = openpyxl.load_workbook(path)
    sheet = wb['Sheet']
    col_num = 1

    for i in range(0, len(tds)):
        # 找到一个有空格子的列
        while(sheet[get_column_letter(col_num) + str(row_num)].value != None):
            col_num = col_num + 1
        try:
            j = int(tds[i]["colspan"])
        except:
            sheet[get_column_letter(col_num) + str(row_num)] = tds[i].getText()
            try:
                r = int(tds[i]["rowspan"])
            except:
                pass
            else:
                for k in range(row_num + 1, row_num + r - 1):
                    sheet[get_column_letter(col_num) + str(k)] = ' '
        else:
            for k in range(0, j):
                try:
                    r = int(tds[i]["rowspan"])
                except:
                    pass
                else:
                    for p in range(row_num + 1, row_num + r - 1):
                        sheet[get_column_letter(col_num) + str(p)] = ' '
                if(k == j/2):
                    sheet[get_column_letter(col_num) + str(row_num)] = tds[i].getText()
                else:
                    sheet[get_column_letter(col_num) + str(row_num)] = ' '
                col_num = col_num + 1
    wb.save(path)    

""" 传入一个web文件名，生成一个xlsx文件"""
def create_xlsx(path):
    # 分析网页
    webFile = open('html/' + path + '.html', encoding='UTF-8')
    webFileSoup = bs4.BeautifulSoup(webFile.read(), 'lxml')
    elems = webFileSoup.select("tr")

    # 计算地区列的位置 
    location = 0
    for i in range(0, len(elems)):
        try:
            if(elems[i].find_all("td")[0].getText() == '地区'):
                location = i
                break
        except:
            pass

    # 计算数据开始的行数
    theFirst = 7
    for i in range(0, len(elems)):
        try:
            if(elems[i].find_all("td")[0].getText() == '全国合计'):
                theFirst = i
                break
        except:
            pass

    # 先将每个网页放入表格中
    # 为网页创建表格
    wb_this = openpyxl.Workbook()
    sheet_this = wb_this['Sheet']
    wb_this.save('xlsx/' + path + '.xlsx')

    # 然后，再从每个表格中获取对应列的数据
    # 写表头
    # 从地区的属性获取总行数rowspan，然后往后遍历rowspan个tr
    rowspan = int(elems[location].find_all("td")[0]["rowspan"])
    for i in range(location, location + rowspan):
        # 调用函数写行
        write_a_row(elems[i].find_all("td"), i - location + 1, 'xlsx/' + path + '.xlsx')

    # 写数据
    i = 1
    for j in range(theFirst, theFirst + 33):
        # 调用函数写行
        write_a_row(elems[j].find_all("td"), rowspan + i, 'xlsx/' + path + '.xlsx')
        i = i + 1

""" 一键获取同类数据 """
def autoGetDate(title_name, name_list):
    theCol = 0
    theRow = 0
    # 创建一个新的excel文件
    wb_new = openpyxl.Workbook()
    sheet_new = wb_new['Sheet']
    wb_new.save(title_name + '.xlsx')

    # 打开所有记录的excel文件，并获得数据，写入新文件中
    firstTime = 1
    for path in name_list:
        # 打开文件
        wb_old = openpyxl.load_workbook('xlsx/' + path + '.xlsx')
        sheet_old = wb_old['Sheet']

        # 如果是第一次写入数据，写入纵表头
        if(firstTime == 1):
            sheet_new[get_column_letter(1) + str(1)] = '地区'
            for i in range(2, 2 + 33):
                sheet_new[get_column_letter(1) + str(i)] = provinces[i-2]

        # 寻找title_name所在的列
        for r in range(1, 5):
            for c in range(2, 500):
                if(sheet_old[get_column_letter(c) + str(r)].value == title_name):
                    theCol = c
                    break

        # 找到开始的行
        for r in range(1, 10):
            if(sheet_old[get_column_letter(1) + str(r)].value == '全国合计'):
                theRow = r
                break

        # 写入表头
        sheet_new[get_column_letter(firstTime + 1) + str(1)] = path

        # 写入数据
        for i in range(2, 2 + 33):
            sheet_new[get_column_letter(firstTime +1) + str(i)] = sheet_old[get_column_letter(theCol) + str(theRow + i - 2)].value

        # 不是第一次了
        firstTime = firstTime + 1
    wb_new.save(title_name + '.xlsx')

# """ 测试 """
# create_xlsx('2019年1季度民政统计分省数据', 1)